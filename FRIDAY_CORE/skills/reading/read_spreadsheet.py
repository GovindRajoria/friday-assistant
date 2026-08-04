# skills/reading/read_spreadsheet.py
"""CSV and XLSX: a shape summary, column statistics, or one specific lookup.

A spreadsheet is the case where dumping the file into the prompt is actively
wrong. Ten thousand rows will not fit, and even when they do, a local model
asked "what is the total" over raw rows will confabulate a number that looks
plausible. So the arithmetic happens in Python and the model is handed the
result — the same reasoning as the anomaly guard, and the same reasoning behind
the prompt rule that forbids inventing a calculator tool.

Three actions, because they answer three different questions:
  summary  — what is in this file at all: sheets, columns, row count, a sample
  column   — statistics for one column: count, numeric min/max/mean/sum, blanks
  find     — the rows where a column matches a value
"""
import csv

from core.paths import allowed_roots, refusal, resolve_within

MAX_SAMPLE_ROWS = 5
MAX_MATCHES = 20
# Read far enough to be useful, and stop long before memory matters.
MAX_ROWS_SCANNED = 50000


class ReadSpreadsheetSkill:
    def __init__(self):
        self.manifest = {
            "name": "read_spreadsheet",
            "description": (
                "Reads a CSV or Excel file on this computer and returns what is in it. "
                "Parameters: 'path'; 'action' (summary for sheets/columns/row count, "
                "column for statistics on one column, find for matching rows); 'column'; "
                "and 'value' for find. Totals and averages are computed exactly, not "
                "estimated. Use read_document for prose files, not this."
            ),
            "parameters": ["path", "action", "column", "value"],
        }

    def execute(self, params=None):
        params = params or {}
        path_str = params.get("path")
        if not path_str:
            return {"status": "error", "message": "I need the path of a spreadsheet to read."}

        resolved = resolve_within(str(path_str), allowed_roots())
        if resolved is None:
            return refusal(str(path_str))
        if not resolved.is_file():
            return {"status": "error", "message": f"'{resolved}' is not a readable file."}

        suffix = resolved.suffix.lower()
        try:
            if suffix in {".csv", ".tsv", ".txt"}:
                headers, rows = self._read_csv(resolved)
                sheet_note = ""
            elif suffix in {".xlsx", ".xlsm"}:
                headers, rows, sheet_note = self._read_xlsx(resolved, params.get("sheet"))
            else:
                return {
                    "status": "error",
                    "message": f"I read .csv and .xlsx; '{suffix or 'no extension'}' is neither.",
                }
        except ImportError as error:
            return {"status": "error", "message": f"I am missing the library for {suffix}: {error}"}
        except Exception as error:                                    # noqa: BLE001
            return {"status": "error", "message": f"Could not read '{resolved.name}': {error}"}

        if not headers:
            return {"status": "error", "message": f"'{resolved.name}' has no header row I could read."}

        action = str(params.get("action") or "summary").lower()
        if action == "column":
            return self._column(resolved, headers, rows, params.get("column"))
        if action in {"find", "lookup", "filter"}:
            return self._find(resolved, headers, rows, params.get("column"), params.get("value"))
        return self._summary(resolved, headers, rows, sheet_note)

    # ---- readers ----------------------------------------------------------

    def _read_csv(self, path):
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel                      # a single-column file sniffs as nothing
            reader = csv.reader(handle, dialect)
            rows = []
            headers = []
            for index, row in enumerate(reader):
                if index == 0:
                    headers = [str(cell).strip() for cell in row]
                    continue
                rows.append(row)
                if index >= MAX_ROWS_SCANNED:
                    break
        return headers, rows

    def _read_xlsx(self, path, sheet):
        import openpyxl

        # read_only keeps a large workbook off the heap; data_only returns the
        # cached result of a formula rather than the formula text, which is what
        # anyone asking "what is the total" means.
        book = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        try:
            names = book.sheetnames
            worksheet = book[sheet] if sheet and sheet in names else book[names[0]]
            rows_iter = worksheet.iter_rows(values_only=True)
            headers = []
            rows = []
            for index, row in enumerate(rows_iter):
                if index == 0:
                    headers = ["" if cell is None else str(cell).strip() for cell in row]
                    continue
                rows.append(["" if cell is None else cell for cell in row])
                if index >= MAX_ROWS_SCANNED:
                    break
            note = f" (sheet '{worksheet.title}' of {len(names)}: {', '.join(names)})" if len(names) > 1 else ""
            return headers, rows, note
        finally:
            book.close()

    # ---- actions ----------------------------------------------------------

    def _summary(self, path, headers, rows, sheet_note):
        sample_lines = []
        for row in rows[:MAX_SAMPLE_ROWS]:
            pairs = [f"{h}={self._render(v)}" for h, v in zip(headers, row) if str(v).strip() != ""]
            sample_lines.append("  " + ", ".join(pairs))
        sample = "\n".join(sample_lines) or "  (no data rows)"
        return {
            "status": "success",
            "message": (f"{path.name}{sheet_note}: {len(rows)} data row(s), "
                        f"{len(headers)} column(s).\nColumns: {', '.join(headers)}\n"
                        f"First {min(len(rows), MAX_SAMPLE_ROWS)} row(s):\n{sample}"),
            "data": {"columns": headers, "row_count": len(rows)},
        }

    def _column(self, path, headers, rows, column):
        index = self._column_index(headers, column)
        if index is None:
            return self._no_such_column(column, headers)

        values = [row[index] if index < len(row) else "" for row in rows]
        filled = [v for v in values if str(v).strip() != ""]
        numbers = []
        for value in filled:
            try:
                numbers.append(float(str(value).replace(",", "").strip()))
            except (TypeError, ValueError):
                continue

        parts = [f"{len(filled)} value(s) present, {len(values) - len(filled)} blank"]
        if numbers:
            total = sum(numbers)
            parts.append(
                f"{len(numbers)} numeric: min {self._number(min(numbers))}, "
                f"max {self._number(max(numbers))}, mean {self._number(total / len(numbers))}, "
                f"sum {self._number(total)}"
            )
        else:
            distinct = sorted({str(v).strip() for v in filled})
            shown = ", ".join(distinct[:10])
            parts.append(f"{len(distinct)} distinct value(s)" + (f": {shown}" if shown else ""))

        return {
            "status": "success",
            "message": f"{path.name}, column '{headers[index]}' — " + "; ".join(parts) + ".",
            "data": {"column": headers[index], "present": len(filled), "numeric": len(numbers)},
        }

    def _find(self, path, headers, rows, column, value):
        if value is None or str(value).strip() == "":
            return {"status": "error", "message": "I need a value to look for."}
        index = self._column_index(headers, column)
        if index is None:
            return self._no_such_column(column, headers)

        needle = str(value).strip().lower()
        matches = []
        for row in rows:
            cell = str(row[index]).strip().lower() if index < len(row) else ""
            if needle in cell:
                matches.append(row)
            if len(matches) >= MAX_MATCHES:
                break

        if not matches:
            return {
                "status": "success",
                "message": f"No row in {path.name} has '{value}' in column '{headers[index]}'.",
                "data": {"matches": 0},
            }

        lines = []
        for row in matches:
            pairs = [f"{h}={self._render(v)}" for h, v in zip(headers, row) if str(v).strip() != ""]
            lines.append("  " + ", ".join(pairs))
        capped = " (first 20)" if len(matches) >= MAX_MATCHES else ""
        return {
            "status": "success",
            "message": (f"{len(matches)} row(s){capped} in {path.name} where '{headers[index]}' "
                        f"contains '{value}':\n" + "\n".join(lines)),
            "data": {"matches": len(matches)},
        }

    # ---- helpers ----------------------------------------------------------

    @staticmethod
    def _column_index(headers, column):
        """Match a column by exact name, then case-insensitively, then by position."""
        if column is None or str(column).strip() == "":
            return None
        wanted = str(column).strip()
        if wanted in headers:
            return headers.index(wanted)
        lowered = [h.lower() for h in headers]
        if wanted.lower() in lowered:
            return lowered.index(wanted.lower())
        # The model sometimes answers with a number for "which column".
        if wanted.isdigit() and 1 <= int(wanted) <= len(headers):
            return int(wanted) - 1
        return None

    @staticmethod
    def _no_such_column(column, headers):
        return {
            "status": "error",
            "message": (f"There is no column '{column}'. The columns are: {', '.join(headers)}."),
        }

    @staticmethod
    def _number(value: float) -> str:
        """Whole numbers without a trailing .0; everything else to 2 places."""
        return str(int(value)) if float(value).is_integer() else f"{value:.2f}"

    @staticmethod
    def _render(value) -> str:
        text = str(value)
        return text if len(text) <= 40 else text[:37] + "..."


def setup():
    return ReadSpreadsheetSkill()
